import json
import os
import decimal
import tempfile
from datetime import datetime

from PyQt6.QtWidgets import (
    QWidget, QDialog, QVBoxLayout, QHBoxLayout, QGridLayout, QGroupBox,
    QLabel, QLineEdit, QPushButton, QComboBox, QCheckBox, QSpinBox,
    QDoubleSpinBox, QTreeWidget, QTreeWidgetItem, QFileDialog, QMessageBox,
    QRadioButton, QHeaderView,
)
from PyQt6.QtCore import Qt, QRect, pyqtSignal
from PyQt6.QtGui import QFont, QColor, QPainter, QPen

from db.bom_loader import BOMLoader


def _write_indicator_svgs() -> tuple[str, str]:
    """
    Write checkbox indicator SVGs to the system temp folder and return
    (unchecked_url, checked_url) ready for use in a Qt stylesheet url().
    Written every run so changes survive temp-folder cleans.
    """
    unchecked = (
        '<svg width="13" height="13" xmlns="http://www.w3.org/2000/svg">'
        '<rect x="1" y="1" width="11" height="11" rx="2"'
        ' fill="white" stroke="#888888" stroke-width="1.5"/>'
        '</svg>'
    )
    checked = (
        '<svg width="13" height="13" xmlns="http://www.w3.org/2000/svg">'
        '<rect x="1" y="1" width="11" height="11" rx="2"'
        ' fill="white" stroke="#1565C0" stroke-width="1.5"/>'
        '<polyline points="3,7 5.5,9.5 10.5,4" fill="none"'
        ' stroke="#1565C0" stroke-width="2"'
        ' stroke-linecap="round" stroke-linejoin="round"/>'
        '</svg>'
    )
    tmp = tempfile.gettempdir()
    for fname, content in [('bom_chk_off.svg', unchecked),
                            ('bom_chk_on.svg',  checked)]:
        with open(os.path.join(tmp, fname), 'w', encoding='utf-8') as fh:
            fh.write(content)
    # Qt stylesheet urls need forward slashes even on Windows
    off = os.path.join(tmp, 'bom_chk_off.svg').replace('\\', '/')
    on  = os.path.join(tmp, 'bom_chk_on.svg').replace('\\', '/')
    return off, on


class _JsonEncoder(json.JSONEncoder):
    """Handles Decimal and any other non-standard types pyodbc may return."""
    def default(self, obj):
        if isinstance(obj, decimal.Decimal):
            return float(obj)
        return super().default(obj)

# Sentinel stored in UserRole to mark placeholder children
_PLACEHOLDER = '__placeholder__'

# UserRole slots
_ROLE_ITEM_NO  = Qt.ItemDataRole.UserRole          # 256 — stores item number string
_ROLE_HAS_BOM  = Qt.ItemDataRole.UserRole + 1      # 257 — stores bool (BILLTYPE == 1)

# Columns that should sort numerically (Position=0, Qty=2)
_NUMERIC_SORT_COLS = frozenset({0, 2, 7})


def _fmt_qty(val) -> str:
    """Format a qty value: whole numbers show as int (e.g. 1.0 → '1'),
    fractional values keep their decimal (e.g. 1.5 → '1.5')."""
    if val is None or val == '':
        return ''
    try:
        f = float(val)
        return str(int(f)) if f == int(f) else str(f)
    except (ValueError, TypeError):
        return str(val)


class _BOMTreeItem(QTreeWidgetItem):
    """QTreeWidgetItem with column-aware sorting.

    - Position (col 0) and Qty (col 2) sort numerically.
    - All other columns sort case-insensitively.
    - Placeholder loading items always sort to the bottom.
    """

    def __lt__(self, other: QTreeWidgetItem) -> bool:  # noqa: D105
        # Placeholders always go to the bottom regardless of sort direction
        self_is_ph  = self.data(0, _ROLE_ITEM_NO) == _PLACEHOLDER
        other_is_ph = other.data(0, _ROLE_ITEM_NO) == _PLACEHOLDER
        if self_is_ph and other_is_ph:
            return False
        if self_is_ph:
            return False   # self sinks to bottom
        if other_is_ph:
            return True    # other sinks to bottom

        tree = self.treeWidget()
        col  = tree.sortColumn() if tree else 1

        a = self.text(col)
        b = other.text(col)

        if col in _NUMERIC_SORT_COLS:
            try:
                return float(a or 0) < float(b or 0)
            except (ValueError, TypeError):
                pass
        return a.casefold() < b.casefold()

# Background colors by BOM depth (index = depth, clamped at last entry)
_DEPTH_COLORS = [
    '#FFFFFF',  # depth 0 — root
    '#E6E6E6',  # depth 1  (Δ25)
    '#CCCCCC',  # depth 2  (Δ26)
    '#B3B3B3',  # depth 3  (Δ25)
    '#999999',  # depth 4  (Δ26)
    '#808080',  # depth 5  (Δ25)
    '#666666',  # depth 6+ (Δ26)
]



class CheckableHeaderView(QHeaderView):
    """QHeaderView that paints a checkbox at the left edge of each section.

    Clicking the checkbox toggles the column's check state and emits
    `checkStateChanged(logicalIndex, checked)`. Clicking elsewhere on the
    header still triggers the normal sort behavior.
    """

    checkStateChanged = pyqtSignal(int, bool)

    _CHK_SIZE = 13
    _CHK_PAD  = 4

    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self._checks: dict[int, bool] = {}
        self.setSectionsClickable(True)

    def set_checked(self, col: int, checked: bool, *, emit: bool = False):
        prev = self._checks.get(col, True)
        self._checks[col] = checked
        if prev != checked:
            self.updateSection(col)
            if emit:
                self.checkStateChanged.emit(col, checked)

    def is_checked(self, col: int) -> bool:
        return self._checks.get(col, True)

    def _checkbox_rect(self, section_x: int, section_w: int) -> QRect:
        h = self.height()
        x = section_x + self._CHK_PAD
        y = (h - self._CHK_SIZE) // 2
        return QRect(x, y, self._CHK_SIZE, self._CHK_SIZE)

    def paintSection(self, painter, rect, logicalIndex):
        # Reserve space on the left for the checkbox by shifting the default
        # label/sort indicator rendering to the right.
        offset = self._CHK_SIZE + self._CHK_PAD * 2
        shifted = QRect(rect.x() + offset, rect.y(),
                        max(0, rect.width() - offset), rect.height())
        super().paintSection(painter, shifted, logicalIndex)

        # Re-paint the section background slice that lives under the checkbox
        # so the default header gradient shows through there as well.
        bg_rect = QRect(rect.x(), rect.y(), offset, rect.height())
        opt = self.style().standardPalette()
        painter.save()
        painter.fillRect(bg_rect, opt.button())

        chk_rect = QRect(rect.x() + self._CHK_PAD,
                         rect.y() + (rect.height() - self._CHK_SIZE) // 2,
                         self._CHK_SIZE, self._CHK_SIZE)
        checked = self.is_checked(logicalIndex)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.setBrush(QColor('#FFFFFF'))
        border = QColor('#1565C0') if checked else QColor('#888888')
        pen = QPen(border)
        pen.setWidthF(1.5)
        painter.setPen(pen)
        painter.drawRoundedRect(chk_rect, 2, 2)
        if checked:
            tick = QPen(QColor('#1565C0'))
            tick.setWidthF(2.0)
            tick.setCapStyle(Qt.PenCapStyle.RoundCap)
            tick.setJoinStyle(Qt.PenJoinStyle.RoundJoin)
            painter.setPen(tick)
            r = chk_rect
            painter.drawLine(r.x() + 3, r.y() + 7,
                             r.x() + 5, r.y() + 9)
            painter.drawLine(r.x() + 5, r.y() + 9,
                             r.x() + 10, r.y() + 4)
        painter.restore()

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            pos = event.pos()
            section = self.logicalIndexAt(pos)
            if section >= 0:
                section_x = self.sectionViewportPosition(section)
                hit = QRect(section_x, 0,
                            self._CHK_SIZE + self._CHK_PAD * 2, self.height())
                if hit.contains(pos):
                    new_state = not self.is_checked(section)
                    self.set_checked(section, new_state, emit=True)
                    return  # swallow click — don't trigger sort
        super().mousePressEvent(event)


_DEPTH_COLOR_pdf = [
    '#FFFFFF',  # depth 0 — root
    '#FFFFFF',  # depth 1  (Δ25)
    '#FFFFFF',  # depth 2  (Δ26)
    '#FFFFFF',  # depth 3  (Δ25)
    '#FFFFFF',  # depth 4  (Δ26)
    '#FFFFFF',  # depth 5  (Δ25)
    '#FFFFFF',  # depth 6+ (Δ26)
]
# Column definitions for PDF settings dialog
# (label, default_with_pos, default_without_pos)  — None = not shown
# Order: left-side cols first (Position, Item No.), then right-side cols (Qty, Drawing).
# Description width is auto-computed and not listed here.
_PDF_COL_DEFS = [
    ('Position',                    2.0,  None),
    ('Artikel-Nr./Item\u00a0No.',   3.8,  5.0),
    ('Stück\u00a0/\u00a0Qty',       1.0,  1.5),
    ('Draw No.',                 1.8,  2.5),
]


class PDFSettingsDialog(QDialog):
    """Lets the user tweak font sizes and column widths before exporting/previewing.

    `cols` is a list of (label, default_width_cm) for the columns the user
    chose to include via the BOM tree's header checkboxes.
    """

    def __init__(self, cols: list, preview_callback, parent=None):
        super().__init__(parent)
        self.setWindowTitle("PDF Export Settings")
        self.setModal(True)
        self.setMinimumWidth(460)
        self._preview_cb  = preview_callback
        self._cols = list(cols)
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # ── Font sizes ────────────────────────────────────────────────
        font_box = QGroupBox("Font Sizes")
        fl = QHBoxLayout(font_box)
        fl.addWidget(QLabel("Header:"))
        self._hdr_fs = QSpinBox()
        self._hdr_fs.setRange(5, 18); self._hdr_fs.setValue(12); self._hdr_fs.setSuffix(" pt")
        self._hdr_fs.setFixedWidth(72)
        fl.addWidget(self._hdr_fs)
        fl.addSpacing(28)
        fl.addWidget(QLabel("Body:"))
        self._body_fs = QSpinBox()
        self._body_fs.setRange(5, 16); self._body_fs.setValue(11); self._body_fs.setSuffix(" pt")
        self._body_fs.setFixedWidth(72)
        fl.addWidget(self._body_fs)
        fl.addStretch()
        layout.addWidget(font_box)

        # ── Page Orientation ─────────────────────────────────────
        orient_box = QGroupBox("Page Orientation")
        ol = QHBoxLayout(orient_box)

        self._portrait = QRadioButton("Vertical (Portrait)")
        self._landscape = QRadioButton("Horizontal (Landscape)")

        self._portrait.setChecked(True)

        ol.addWidget(self._portrait)
        ol.addWidget(self._landscape)
        ol.addStretch()

        layout.addWidget(orient_box)

        # ── Column widths ─────────────────────────────────────────────
        col_box = QGroupBox("Column Widths (cm)")
        gl = QGridLayout(col_box)
        gl.setHorizontalSpacing(6)
        gl.setVerticalSpacing(6)
        gl.setColumnMinimumWidth(3, 18)   # gap between the two halves
        self._col_spins: dict[str, QDoubleSpinBox] = {}
        for i, (label, default) in enumerate(self._cols):
            r, c = divmod(i, 2)
            base = c * 4   # cols: 0/4=label, 1/5=spin, 2/6=unit, 3=gap
            gl.addWidget(QLabel(f"{label}:"), r, base)
            spin = QDoubleSpinBox()
            spin.setRange(0.5, 20.0); spin.setSingleStep(0.1)
            spin.setDecimals(1); spin.setValue(default); spin.setFixedWidth(74)
            self._col_spins[label] = spin
            gl.addWidget(spin, r, base + 1)
            gl.addWidget(QLabel("cm"), r, base + 2)
        layout.addWidget(col_box)

        # ── Buttons ───────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_prev = QPushButton("Preview")
        btn_prev.setToolTip("Generate a temporary PDF and open it in your PDF viewer")
        btn_prev.clicked.connect(self._on_preview)
        btn_row.addWidget(btn_prev)
        btn_row.addStretch()
        btn_cancel = QPushButton("Cancel"); btn_cancel.clicked.connect(self.reject)
        btn_export = QPushButton("Export"); btn_export.setDefault(True)
        btn_export.clicked.connect(self.accept)
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_export)
        layout.addLayout(btn_row)

    def settings(self) -> dict:
        orientation = "portrait"
        if self._landscape.isChecked():
            orientation = "landscape"
        return {
            'header_font_size': self._hdr_fs.value(),
            'body_font_size':   self._body_fs.value(),
            "orientation": orientation,
            'col_widths':       [self._col_spins[lbl].value() for lbl, _ in self._cols],
        }

    def _on_preview(self):
        self._preview_cb(self.settings())


class BOMPanel(QWidget):
    """
    BOM Tree tab.

    How lazy loading works:
    ┌─────────────────────────────────────────────────────┐
    │ 1. User enters item number → Load BOM button        │
    │ 2. Root node created, BOMLoader fires for it        │
    │ 3. Children added; if BILLTYPE==1 → placeholder     │
    │    child added so Qt shows the ▶ expand arrow       │
    │ 4. User clicks ▶  → itemExpanded fires              │
    │ 5. Placeholder removed, BOMLoader fires for that    │
    │    child → its children are added, and so on...     │
    └─────────────────────────────────────────────────────┘

    Export reads the current tree widget state — expanded nodes include
    their children, collapsed nodes do not.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self._active_loaders = []   # keep refs so GC won't kill running threads
        self._export_pending = False
        self._expand_all_active = False  # while True, newly loaded children auto-expand
        self._setup_ui()

    # ------------------------------------------------------------------ setup
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(6)

        # --- Top bar ---
        top = QHBoxLayout()
        self._item_input = QLineEdit()
        self._item_input.setPlaceholderText("Enter item number  e.g. 7956271.00")
        self._item_input.returnPressed.connect(self._load_root)

        self._dataset_cb = QComboBox()
        self._dataset_cb.addItems(['INL', 'KON'])
        self._dataset_cb.setFixedWidth(70)

        self._btn_load = QPushButton("Load BOM")
        self._btn_load.setFixedWidth(100)
        self._btn_load.clicked.connect(self._load_root)

        self._btn_clear = QPushButton("Clear")
        self._btn_clear.setFixedWidth(60)
        self._btn_clear.clicked.connect(self._clear)

        # Export format selector + button
        self._export_fmt = QComboBox()
        self._export_fmt.addItems(["PDF", "JSON", "Excel" ])
        self._export_fmt.setFixedWidth(68)
        self._export_fmt.setToolTip("Choose export format")

        self._btn_export = QPushButton("Export BOM")
        self._btn_export.setFixedWidth(110)
        self._btn_export.setToolTip(
            "Export the currently visible tree as JSON / Excel / PDF.\n"
            "Expanded nodes include their children; collapsed nodes do not."
        )
        self._btn_export.clicked.connect(self._export_bom)

        self._chk_unique = QCheckBox("Skip duplicate rows")
        self._chk_unique.setChecked(True)
        self._chk_unique.setToolTip(
            "Checked  → hide rows where Position + ItemNo appear more than once\n"
            "Unchecked → show every raw row returned by the query"
        )

        self._btn_expand_all = QPushButton("Expand All")
        self._btn_expand_all.setFixedWidth(90)
        self._btn_expand_all.setToolTip(
            "Expand the entire BOM tree, loading every sub-level on demand.\n"
            "Newly loaded children are also expanded automatically."
        )
        self._btn_expand_all.clicked.connect(self._on_expand_all)

        self._btn_collapse_all = QPushButton("Collapse All")
        self._btn_collapse_all.setFixedWidth(95)
        self._btn_collapse_all.setToolTip("Collapse every node in the BOM tree.")
        self._btn_collapse_all.clicked.connect(self._on_collapse_all)

        top.addWidget(QLabel("Item No:"))
        top.addWidget(self._item_input, 1)
        top.addWidget(QLabel("Dataset:"))
        top.addWidget(self._dataset_cb)
        top.addWidget(self._btn_load)
        top.addWidget(self._btn_clear)
        layout.addLayout(top)

        # --- Export bar ---
        export_bar = QHBoxLayout()
        export_bar.addWidget(self._chk_unique)
        export_bar.addSpacing(12)
        export_bar.addWidget(self._btn_expand_all)
        export_bar.addWidget(self._btn_collapse_all)
        export_bar.addStretch()
        export_bar.addWidget(QLabel("Export:"))
        export_bar.addWidget(self._export_fmt)
        export_bar.addWidget(self._btn_export)
        layout.addLayout(export_bar)

        # --- Status ---
        self._status = QLabel("Enter an item number and click Load BOM.")
        layout.addWidget(self._status)

        # --- Tree ---
        self._tree = QTreeWidget()
        self._tree.setHeaderLabels([
            'Position', 'Item No', 'Qty', 'SCRIPTNUM', 'Description', 'Full Name',
            'Lagerort', 'Bestand'
        ])
        _chk_off, _chk_on = _write_indicator_svgs()
        self._tree.setStyleSheet(f"""
            QTreeWidget::item:selected {{
                background-color: #1565C0;
                color: white;
            }}
            QTreeWidget::indicator:unchecked {{
                image: url({_chk_off});
                width: 13px;
                height: 13px;
            }}
            QTreeWidget::indicator:checked {{
                image: url({_chk_on});
                width: 13px;
                height: 13px;
            }}
        """)
        self._tree.setColumnWidth(0, 90)
        self._tree.setColumnWidth(1, 160)
        self._tree.setColumnWidth(2, 65)
        self._tree.setColumnWidth(3, 75)
        self._tree.setColumnWidth(4, 220)
        self._tree.setColumnWidth(5, 280)
        self._tree.setColumnWidth(6, 90)
        self._tree.setColumnWidth(7, 80)
        self._tree.setAlternatingRowColors(False)   # depth shading replaces this
        self._tree.setUniformRowHeights(True)
        self._tree.itemExpanded.connect(self._on_item_expanded)
        self._tree.itemChanged.connect(self._on_item_check_changed)

        # ── Column-header sorting with checkboxes ─────────────────────
        # Replace standard header with checkable version
        old_hdr = self._tree.header()
        old_hdr.deleteLater()
        hdr = CheckableHeaderView(Qt.Orientation.Horizontal, self._tree)
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setMinimumSectionSize(90)
        hdr.setSectionsClickable(True)
        hdr.setSortIndicatorShown(True)
        hdr.setToolTip(
            "Click a column header to sort A→Z.\n"
            "Click again to reverse (Z→A).\n"
            "Click checkbox to include/exclude column in PDF export."
        )
        hdr.sortIndicatorChanged.connect(self._on_sort_indicator_changed)
        hdr.checkStateChanged.connect(self._on_header_check_changed)
        self._tree.setHeader(hdr)

        # Initialize PDF-export check state per column.
        # Position is OFF by default (gives Designation more room); the rest are ON.
        _PDF_DEFAULT_CHECKED = {0: False, 1: True, 2: True, 3: True, 4: True, 5: True, 6: True, 7: True}
        for i in range(self._tree.columnCount()):
            hdr.set_checked(i, _PDF_DEFAULT_CHECKED.get(i, True))

        layout.addWidget(self._tree)

    # ------------------------------------------------------------------ public
    def load_item(self, item_no: str):
        """Can be called externally (e.g. from table explorer double-click)."""
        self._item_input.setText(item_no)
        self._load_root()

    # ------------------------------------------------------------------ slots
    def _load_root(self):
        item_no = self._item_input.text().strip()
        if not item_no:
            return

        self._tree.clear()
        self._active_loaders.clear()
        # self._export_pending = False
        self._status.setText(f"Loading BOM for  {item_no} ...")

        # Root node — has_bom=False so NO placeholder is added.
        root = self._make_node(
            parent=self._tree,
            pos='', item_no=item_no,
            qty='', has_bom=False,
            description='Loading...', full_name='', scriptnum=''
        )
        root.setExpanded(True)

        bold = QFont()
        bold.setBold(True)
        for col in range(self._tree.columnCount()):
            root.setFont(col, bold)

        self._start_loader(item_no, root)

    def _on_sort_indicator_changed(self, col: int, order: Qt.SortOrder):
        """Qt has already toggled the indicator — just apply the sort."""
        self._tree.sortItems(col, order)

    def _on_header_check_changed(self, col: int, checked: bool):
        """Column header checkbox toggled. State is stored in the header
        view itself; PDF export reads it at export time, so nothing else
        needs to happen here. Hook is kept for future use / extension."""
        return

    def _on_expand_all(self):
        """Expand every node, lazily loading and re-expanding as data arrives."""
        if self._tree.invisibleRootItem().childCount() == 0:
            return
        self._expand_all_active = True
        self._tree.expandAll()

    def _on_collapse_all(self):
        """Collapse back to the initial post-load view: root expanded with its
        direct children visible, every deeper level collapsed."""
        self._expand_all_active = False
        self._tree.collapseAll()
        root = self._tree.invisibleRootItem().child(0)
        if root is not None:
            root.setExpanded(True)

    def _on_item_expanded(self, item: QTreeWidgetItem):
        """Fired when user clicks ▶ on a tree node."""
        if item.childCount() != 1:
            return
        placeholder = item.child(0)
        if placeholder.data(0, _ROLE_ITEM_NO) != _PLACEHOLDER:
            return

        # Remove placeholder and load real children
        item_no = item.data(0, _ROLE_ITEM_NO)
        item.removeChild(placeholder)
        self._status.setText(f"Loading children for  {item_no} ...")
        self._start_loader(item_no, item)

    def _on_data_ready(self, parent_item: QTreeWidgetItem, rows: list):
        if not rows:
            self._status.setText(
                f"No BOM found for  {parent_item.data(0, _ROLE_ITEM_NO)}"
            )
            parent_item.setData(0, _ROLE_HAS_BOM, False)
            return

        self._tree.blockSignals(True)
        try:
            # Update parent node label with father info from first row
            first = rows[0]
            parent_item.setText(4, str(first.get('FatherDescription') or ''))
            parent_item.setText(5, str(first.get('FatherFullName')    or ''))

            # Re-apply blue foreground — setText() can silently reset the color
            if parent_item.data(0, _ROLE_HAS_BOM):
                for col in range(self._tree.columnCount()):
                    parent_item.setForeground(col, QColor('#1565C0'))

            # ── Deduplicate by (Position, ItemNo) if checkbox is checked ──
            if self._chk_unique.isChecked():
                seen = set()
                unique_rows = []
                for row in rows:
                    key = (str(row.get('Position') or ''), str(row.get('ItemNo') or ''))
                    if key not in seen:
                        seen.add(key)
                        unique_rows.append(row)
                dup_count = len(rows) - len(unique_rows)
                rows = unique_rows
            else:
                dup_count = 0

            for row in rows:
                has_bom = (row.get('Artikelart') == 1)

                child = self._make_node(
                    parent=parent_item,
                    pos=str(row.get('Position')    or ''),
                    item_no=str(row.get('ItemNo')  or ''),
                    qty=_fmt_qty(row.get('Qty')),
                    has_bom=has_bom,
                    description=str(row.get('Description') or ''),
                    full_name=str(row.get('FullName')       or ''),
                    scriptnum=str(row.get('ScriptNum') or ''),
                    stockloc=str(row.get('StockLoc') or ''),
                    bestand=row.get('Bestand', ''),
                )

                if has_bom:
                    for col in range(self._tree.columnCount()):
                        child.setForeground(col, QColor('#1565C0'))  # blue

            # ── Restore expanded state ────────────────────────────────────
            # Qt resets isExpanded() to False when the placeholder child is
            # removed (childCount drops to 0).  Re-expand now that real
            # children exist, so _widget_item_to_dict will recurse into them.
            parent_item.setExpanded(True)

            # ── Inherit parent's check state for newly loaded children ──
            if parent_item.checkState(0) == Qt.CheckState.Unchecked:
                self._cascade_check(parent_item, Qt.CheckState.Unchecked)

        finally:
            self._tree.blockSignals(False)

        # If "Expand All" is active, expand every newly loaded child that has
        # a sub-BOM and directly kick off its loader. We do not rely on the
        # itemExpanded signal here — for freshly created items inside an async
        # callback chain, the signal cascade can be missed, leaving deep levels
        # unloaded. Calling _start_loader explicitly guarantees every level
        # gets fetched.
        if self._expand_all_active:
            for i in range(parent_item.childCount()):
                child = parent_item.child(i)
                if not child.data(0, _ROLE_HAS_BOM):
                    continue
                child.setExpanded(True)
                if (child.childCount() == 1
                        and child.child(0).data(0, _ROLE_ITEM_NO) == _PLACEHOLDER):
                    child_item_no = child.data(0, _ROLE_ITEM_NO)
                    child.removeChild(child.child(0))
                    self._start_loader(child_item_no, child)

        dup_str = f"  ({dup_count} duplicate(s) hidden)" if dup_count else ""
        self._status.setText(
            f"{parent_item.data(0, _ROLE_ITEM_NO)}  —  "
            f"{len(rows)} child item(s) loaded{dup_str}"
        )

    def _on_error(self, msg: str):
        self._status.setText(f"Error: {msg}")

    # ------------------------------------------------------------------ export
    def _export_bom(self):
        if self._tree.invisibleRootItem().childCount() == 0:
            QMessageBox.information(self, "Export", "Load a BOM first.")
            return

        # Expand any unloaded (checked) placeholder nodes before exporting
        root = self._tree.invisibleRootItem().child(0)
        fired = self._expand_all_placeholders(root) if root else 0

        if fired > 0 or self._active_loaders:
            self._export_pending = True
            self._status.setText("Loading unread nodes before export…")
            return

        self._do_export()

    def _expand_all_placeholders(self, item: QTreeWidgetItem) -> int:
        """Recursively fire loaders for expanded+checked nodes whose children are not yet loaded.
        Skips nodes the user never opened (not expanded). Returns the number of loaders started."""
        if item is None:
            return 0
        # Skip unchecked nodes — they won't appear in the export anyway
        if item.checkState(0) == Qt.CheckState.Unchecked:
            return 0
        # Skip nodes the user never expanded — don't load their children for export
        if not item.isExpanded():
            return 0

        fired = 0
        if item.childCount() == 1 and item.child(0).data(0, _ROLE_ITEM_NO) == _PLACEHOLDER:
            item_no = item.data(0, _ROLE_ITEM_NO)
            item.removeChild(item.child(0))
            self._start_loader(item_no, item)
            fired += 1
        else:
            for i in range(item.childCount()):
                fired += self._expand_all_placeholders(item.child(i))
        return fired

    def _do_export(self):
        """Dispatch to the correct save handler once all nodes are loaded."""
        fmt  = self._export_fmt.currentText()
        data = self._build_export_data_from_tree()

        if fmt == 'JSON':
            path, _ = QFileDialog.getSaveFileName(
                self, "Save BOM as JSON",
                f"BOM_{data['metadata']['item_no']}.json",
                "JSON Files (*.json)",
            )
            if not path:
                return
            try:
                self._save_as_json(data, path)
                self._status.setText(f"Saved {data['metadata']['total_items']} items → {path}")
            except Exception as e:
                QMessageBox.critical(self, "Export Error", str(e))
                self._status.setText(f"Export error: {e}")

        elif fmt == 'Excel':
            path, _ = QFileDialog.getSaveFileName(
                self, "Save BOM as Excel",
                f"BOM_{data['metadata']['item_no']}.xlsx",
                "Excel Files (*.xlsx)",
            )
            if not path:
                return
            try:
                self._save_as_excel(data, path)
                self._status.setText(f"Saved {data['metadata']['total_items']} items → {path}")
            except Exception as e:
                QMessageBox.critical(self, "Export Error", str(e))
                self._status.setText(f"Export error: {e}")

        elif fmt == 'PDF':
            self._export_as_pdf(data)



    def _active_pdf_cols(self) -> list[tuple[str, float, str]]:
        """Active left/right PDF columns based on tree-header checkbox state.

        Returns (dialog_label, default_width_cm, key) tuples in left→right order.
        Designation (description / full_name) is laid out automatically and is
        not part of this list.
        """
        hdr = self._tree.header()
        show_pos  = hdr.is_checked(0)
        show_item = hdr.is_checked(1)
        show_qty  = hdr.is_checked(2)
        show_drw  = hdr.is_checked(3)
        show_loc  = hdr.is_checked(6)
        show_stk  = hdr.is_checked(7)
        cols: list[tuple[str, float, str]] = []
        if show_pos:
            cols.append(('Position', 2.0, 'pos'))
        if show_item:
            cols.append(('Artikel-Nr./Item No.',
                         3.8 if show_pos else 5.0, 'item'))
        if show_qty:
            cols.append(('Stück / Qty',
                         1.0 if show_pos else 1.5, 'qty'))
        if show_drw:
            cols.append(('Draw No.', 1.8 if show_pos else 2.5, 'drw'))
        if show_loc:
            cols.append(('Lagerort', 2.0, 'stockloc'))
        if show_stk:
            cols.append(('Bestand', 1.8, 'bestand'))
        return cols

    def _export_as_pdf(self, data: dict):
        """Show PDF settings dialog (with preview), then save."""
        active_cols = self._active_pdf_cols()
        dialog_cols = [(lbl, w) for (lbl, w, _key) in active_cols]

        def _preview(settings: dict):
            fd, tmp_path = tempfile.mkstemp(suffix='.pdf', prefix='bom_preview_')
            os.close(fd)
            try:
                self._save_as_pdf(data, tmp_path, settings)
                os.startfile(tmp_path)
            except Exception as e:
                QMessageBox.critical(self, "Preview Error", str(e))

        dlg = PDFSettingsDialog(dialog_cols, _preview, parent=self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        settings = dlg.settings()
        path, _ = QFileDialog.getSaveFileName(
            self, "Save BOM as PDF",
            f"BOM_{data['metadata']['item_no']}.pdf",
            "PDF Files (*.pdf)",
        )
        if not path:
            return

        try:
            self._save_as_pdf(data, path, settings)
            self._status.setText(f"Saved {data['metadata']['total_items']} items → {path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))
            self._status.setText(f"Export error: {e}")

    # ------------------------------------------------------------------ tree → data
    def _build_export_data_from_tree(self) -> dict:
        """Read the current QTreeWidget and build the export data dict."""
        root_qt = self._tree.invisibleRootItem().child(0)
        bom     = self._widget_item_to_dict(root_qt, is_root=True) if root_qt else {}
        if bom is None:
            bom = {}
        # Use the flattened row count so "Items: N" matches what appears in
        # the exported table (negative-qty rows are excluded by _flatten_bom).
        total   = len(self._flatten_bom(bom))
        return {
            'metadata': {
                'item_no': self._item_input.text().strip(),
                'description': bom.get('description', ''),
                'full_name': bom.get('full_name', ''),
                'dataset': self._dataset_cb.currentText(),
                'exported_at': datetime.now().isoformat(timespec='seconds'),
                'total_items': total,
            },
            'bom': bom,
        }

    def _widget_item_to_dict(self, item: QTreeWidgetItem, is_root: bool = False):
        if item.checkState(0) == Qt.CheckState.Unchecked:
            return None

        qty_text = item.text(2)
        try:
            qty = float(qty_text) if qty_text else ''
        except ValueError:
            qty = qty_text

        node = {
            'position': '' if is_root else item.text(0),
            'item_no': item.text(1),
            'qty': qty,
            'scriptnum': item.text(3),
            'has_bom': item.data(0, _ROLE_HAS_BOM) or False,
            'description': item.text(4),
            'full_name': item.text(5),
            'stockloc': item.text(6),
            'bestand': item.text(7),
            'children': [],
        }

        # Only recurse into children if the node is currently expanded.
        # A node that was opened then closed should NOT include its sub-children.
        if item.isExpanded():
            for i in range(item.childCount()):
                child = item.child(i)
                if child.data(0, _ROLE_ITEM_NO) == _PLACEHOLDER:
                    continue
                child_dict = self._widget_item_to_dict(child)
                if child_dict is not None:
                    node['children'].append(child_dict)

        return node

    def _count_nodes(self, node: dict) -> int:
        if not node:
            return 0
        return 1 + sum(self._count_nodes(c) for c in node.get('children', []))

    # ------------------------------------------------------------------ save helpers
    def _save_as_json(self, data: dict, path: str):
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, cls=_JsonEncoder, indent=2, ensure_ascii=False)

    def _flatten_bom(self, node: dict, level: int = 0, parent_position: str = '',
                     _depth_last: list = None) -> list:
        """Recursively flatten nested BOM dict into a list of row dicts.
        Rows with negative qty are excluded (along with their children).
        _depth_last: list of bools — True means that depth-level node was the last sibling.
        """
        if _depth_last is None:
            _depth_last = []

        qty = node.get('qty', '')
        if isinstance(qty, (int, float)) and qty < 0:
            return []

        position = str(node.get('position') or '').strip()
        item_no = str(node.get('item_no') or '').strip()

        # Build hierarchical position path, e.g. 010.020.030
        if parent_position and position:
            position_path = f"{parent_position}.{position}"
        else:
            position_path = position

        # Pre-filter negative-qty children BEFORE building row dict
        # (visible_children is referenced in the row dict below)
        visible_children = [
            c for c in node.get('children', [])
            if not (isinstance(c.get('qty', ''), (int, float)) and c.get('qty', '') < 0)
        ]

        row = {
            'level': level,
            'position': position,
            'position_path': position_path,
            'item_no': item_no,
            'qty': qty,
            'scriptnum': str(node.get('scriptnum') or '').strip(),
            'description': str(node.get('description') or '').strip(),
            'full_name': str(node.get('full_name') or '').strip(),
            'stockloc': str(node.get('stockloc') or '').strip(),
            'bestand': str(node.get('bestand') or '').strip(),
            '_has_bom_raw': bool(visible_children),
            '_depth_last': list(_depth_last),
        }

        result = [row]

        for idx, child in enumerate(visible_children):
            is_last = (idx == len(visible_children) - 1)
            result.extend(self._flatten_bom(child, level + 1, position_path,
                                            _depth_last + [is_last]))

        return result

    def _save_as_excel(self, data: dict, path: str):
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        meta = data.get('metadata', {})
        rows = self._flatten_bom(data.get('bom', {}))
        if rows and rows[0].get("level") == 0:
            rows = rows[1:]

        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"

        # ── Metadata header rows ──────────────────────────────────────
        ws.append([f"BOM Export — {meta.get('item_no', '')}"])
        ws.append([
            f"{meta.get('description', '')}  |  "
            f"Exported: {meta.get('exported_at', '')}"
        ])
        ws.append([])   # blank spacer

        # ── Column header row ─────────────────────────────────────────
        col_headers = ['Level', 'Position', 'Item No', 'Qty', 'SCRIPTNUM', 'Description', 'Full Name', 'Lagerort', 'Bestand']
        ws.append(col_headers)
        hdr_row = ws.max_row
        hdr_fill = PatternFill(fill_type='solid', fgColor='1565C0')
        hdr_font = Font(bold=True, color='FFFFFF')
        for cell in ws[hdr_row]:
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal='center')

        # ── Data rows ─────────────────────────────────────────────────
        thin = Side(style='thin', color='555555')
        grid = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in rows:
            depth   = row['level']
            hex_col = _DEPTH_COLORS[min(depth, len(_DEPTH_COLORS) - 1)].lstrip('#')
            fill    = PatternFill(fill_type='solid', fgColor=hex_col)

            ws.append([
                depth, row['position'], row['item_no'],
                row['qty'], row['scriptnum'],
                row['description'], row['full_name'],
                row['stockloc'], row['bestand'],
            ])

            r = ws.max_row
            for cell in ws[r]:
                cell.fill   = fill
                cell.border = grid
                if depth == 0:
                    cell.font = Font(bold=True)

        # ── Column widths ─────────────────────────────────────────────
        for col_idx, width in enumerate([8, 12, 28, 10, 10, 45, 38, 12, 12], start=1):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width

        wb.save(path)

    def _save_as_pdf(self, data: dict, path: str, settings: dict = None):
        from reportlab.lib.pagesizes import A4, landscape, portrait
        from reportlab.lib import colors
        from reportlab.lib.units import mm, cm
        from reportlab.pdfgen.canvas import Canvas
        from reportlab.pdfbase.pdfmetrics import stringWidth

        # ── depth-band colors ─────────────────────────────────────────
        _DEPTH_COLOR_pdf = [
            '#FFFFFF',  # depth 0 — root
            '#FFFFFF',  # depth 1
            '#FFFFFF',  # depth 2
            '#FFFFFF',  # depth 3
            '#FFFFFF',  # depth 4
            '#FFFFFF',  # depth 5
            '#FFFFFF',  # depth 6+
        ]

        # ── defaults ──────────────────────────────────────────────────
        if settings is None:
            settings = {
                'header_font_size': 9,
                'body_font_size':   8,
                'col_widths':       None,
                'orientation':      'landscape',
            }

        fs_hdr      = settings.get('header_font_size', 9)
        fs_body     = settings.get('body_font_size',   8)
        orientation = settings.get('orientation',      'landscape')

        page_size = portrait(A4) if orientation == 'portrait' else landscape(A4)
        pw, ph    = page_size

        meta = data.get('metadata', {})
        rows = self._flatten_bom(data.get('bom', {}))
        # Skip root node
        if rows and rows[0].get('level') == 0:
            rows = rows[1:]

        # ── layout ────────────────────────────────────────────────────
        lm      = 15 * mm
        rm      = 15 * mm
        # FIX 1: Reduced row height multiplier from 3.8 → 3.0, min from 10mm → 7mm
        row_h   = max(7 * mm, fs_body * 3.0)
        hdr_h = max(8 * mm, fs_hdr * 3.0)
        title_h = 18 * mm
        bot_m   = 12 * mm

        # ── connector zone carved out of the RIGHT side of Designation ──
        conn_w  = 22 * mm

        # Per-column include/exclude state lives on the tree's checkable header.
        hdr = self._tree.header()
        show_pos  = hdr.is_checked(0)
        show_item = hdr.is_checked(1)
        show_qty  = hdr.is_checked(2)
        show_draw = hdr.is_checked(3)
        show_desc = hdr.is_checked(4)
        show_name = hdr.is_checked(5)
        show_loc  = hdr.is_checked(6)
        show_stk  = hdr.is_checked(7)

        # Build the column list in the *original* visual order:
        #   [Position?] [Item No?] [Designation] [Qty?] [Draw?]
        # Designation is auto-sized (width_cm = 0) and is included whenever
        # Description OR Full Name is checked. Default widths track the
        # original layout (narrower when Position is also shown, wider
        # when Position is omitted) so dialog defaults stay consistent.
        col_configs = []
        if show_pos:
            col_configs.append(('Position', 2.0, 'position_path'))
        if show_item:
            col_configs.append(('Artikel-Nr./Item No.',
                                3.8 if show_pos else 5.0, 'item_no'))
        if show_desc or show_name:
            col_configs.append(('Bezeichnung / Designation', 0, 'description'))
        if show_qty:
            col_configs.append(('Stück / Qty',
                                1.0 if show_pos else 1.5, 'qty'))
        if show_draw:
            col_configs.append(('Draw No.',
                                1.8 if show_pos else 2.5, 'scriptnum'))
        if show_loc:
            col_configs.append(('Lagerort', 2.0, 'stockloc'))
        if show_stk:
            col_configs.append(('Bestand', 1.8, 'bestand'))

        n_cols      = len(col_configs)
        table_right = pw - rm

        # Resolve column widths. raw_cw from the settings dialog only carries
        # entries for fixed-width columns (Designation is auto), in the order
        # the dialog displays them. Walk col_configs and pull the next raw_cw
        # value for each non-auto column.
        raw_cw = settings.get('col_widths')
        cw_vals = []
        auto_idx = -1
        non_auto_seen = 0
        for i, (_lbl, width_cm, _key) in enumerate(col_configs):
            if width_cm == 0:
                auto_idx = i
                cw_vals.append(0)
            else:
                if raw_cw and non_auto_seen < len(raw_cw):
                    cw_vals.append(raw_cw[non_auto_seen] * cm)
                else:
                    cw_vals.append(width_cm * cm)
                non_auto_seen += 1

        if auto_idx >= 0:
            fixed_total = sum(cw_vals[j] for j in range(n_cols) if j != auto_idx)
            remaining   = (table_right - lm) - fixed_total
            cw_vals[auto_idx] = max(remaining, conn_w + 4 * mm)

        xs = []
        x = lm
        for i in range(n_cols):
            xs.append(x)
            x += cw_vals[i]

        # Description column (if present) is assumed to be the last one
        desc_idx = -1
        for i, cfg in enumerate(col_configs):
            if cfg[2] == 'description':
                desc_idx = i
                break
        if desc_idx >= 0:
            x_desc = xs[desc_idx]
            desc_col_right = x_desc + cw_vals[desc_idx]
            desc_text_w = cw_vals[desc_idx] - conn_w - 2 * mm
            x_conn_zone = desc_col_right - conn_w
        else:
            x_desc = xs[-1] + cw_vals[-1] if xs else lm
            desc_col_right = x_desc
            desc_text_w = 0
            x_conn_zone = desc_col_right

        def spine_x(level: int) -> float:
            return (desc_col_right - 4 * mm) - level * 3 * mm

        x_arrow_tip = x_conn_zone + 1.5 * mm

        # ── colors ────────────────────────────────────────────────────
        line_clr   = colors.HexColor('#000000')
        asm_clr    = colors.HexColor('#2d74da')
        hdr_bg_clr = colors.HexColor('#1565C0')
        grid_clr   = colors.HexColor('#555555')

        # ── helpers ───────────────────────────────────────────────────
        def fit_text(text, max_w, font, size):
            if not text:
                return ''
            if stringWidth(text, font, size) <= max_w:
                return text
            while text and stringWidth(text + '\u2026', font, size) > max_w:
                text = text[:-1]
            return (text + '\u2026') if text else ''

        def draw_row_bg(c, row_top, depth):
            hex_col = _DEPTH_COLOR_pdf[min(depth, len(_DEPTH_COLOR_pdf) - 1)]
            c.setFillColor(colors.HexColor(hex_col))
            c.rect(lm, row_top - row_h, table_right - lm, row_h, fill=1, stroke=0)

        def draw_grid_line(c, x1, y1, x2, y2, lw=0.3):
            c.setStrokeColor(grid_clr)
            c.setLineWidth(lw)
            c.line(x1, y1, x2, y2)

        def draw_arrow(c, x_spine, y):
            ah = 1.5 * mm
            x_tip = x_arrow_tip
            if x_spine <= x_tip + ah:
                return
            c.setStrokeColor(asm_clr)
            c.setLineWidth(1.0)
            c.line(x_tip + ah, y, x_spine, y)
            c.line(x_tip, y, x_tip + ah, y + ah * 0.8)
            c.line(x_tip, y, x_tip + ah, y - ah * 0.8)

        def draw_dot(c, x, y, r=2.0, color=None):
            clr = color or line_clr
            c.setFillColor(clr)
            c.setStrokeColor(clr)
            c.circle(x, y, r, fill=1, stroke=0)

        def build_last_child_map(all_rows):
            lc_map = {}
            n = len(all_rows)
            for di in range(n):
                if not all_rows[di].get('_has_bom_raw', False):
                    continue
                depth = all_rows[di]['level']
                last_desc = None
                for dj in range(di + 1, n):
                    cd = all_rows[dj]['level']
                    if cd <= depth:
                        break
                    last_desc = dj
                if last_desc is not None:
                    lc_map[di] = last_desc
            return lc_map

        last_child_map = build_last_child_map(rows)

        def draw_connectors(c, page_rows, centers, page_start_idx):
            n      = len(page_rows)
            p_end  = page_start_idx + n - 1

            c.saveState()

            for gi, lci in last_child_map.items():
                if gi >= page_start_idx:
                    continue
                if lci < page_start_idx:
                    continue

                depth = rows[gi]['level']
                sx    = spine_x(depth)
                clr   = asm_clr if depth == 0 else line_clr

                last_on_page_local = None
                for local_i, gr in enumerate(page_rows):
                    gi_of_row = page_start_idx + local_i
                    if (gr['level'] > depth and gi_of_row <= lci):
                        is_desc = True
                        for mid in range(gi + 1, gi_of_row):
                            if rows[mid]['level'] <= depth:
                                is_desc = False
                                break
                        if is_desc:
                            last_on_page_local = local_i

                if last_on_page_local is None:
                    y_from = centers[0] + row_h * 0.5
                    y_to   = centers[-1] - row_h * 0.5
                else:
                    y_from = centers[0] + row_h * 0.5
                    y_to   = centers[last_on_page_local]

                c.setStrokeColor(clr)
                c.setLineWidth(0.8)
                c.line(sx, y_from, sx, y_to)

            # Pass 1: vertical spines for parents ON this page
            for di, row in enumerate(page_rows):
                if not row.get('_has_bom_raw', False):
                    continue
                depth = row['level']
                sx    = spine_x(depth)
                last_desc_cy = None
                for dj in range(di + 1, n):
                    cdepth = page_rows[dj]['level']
                    if cdepth <= depth:
                        break
                    last_desc_cy = centers[dj]
                if last_desc_cy is not None:
                    clr = asm_clr if depth == 0 else line_clr
                    c.setStrokeColor(clr)
                    c.setLineWidth(0.8)
                    c.line(spine_x(depth), centers[di], sx, last_desc_cy)

            # Pass 2: arrows + dots
            for di, row in enumerate(page_rows):
                depth   = row['level']
                has_bom = row.get('_has_bom_raw', False)
                cy      = centers[di]
                if has_bom:
                    draw_arrow(c, spine_x(depth), cy)
                if depth > 0:
                    draw_dot(c, spine_x(depth - 1), cy,
                             color=asm_clr if has_bom else line_clr)

            c.restoreState()

        # ── pagination ────────────────────────────────────────────────
        usable_first = ph - bot_m - title_h - hdr_h
        usable_other = ph - bot_m - 8 * mm  - hdr_h
        rows_fp = max(1, int(usable_first / row_h))
        rows_op = max(1, int(usable_other / row_h))

        pages = []
        idx = 0
        while idx < len(rows):
            cap = rows_fp if not pages else rows_op
            pages.append((idx, rows[idx: idx + cap]))
            idx += cap
        if not pages:
            pages = [(0, [])]
        n_pages = len(pages)

        # ── draw ──────────────────────────────────────────────────────
        c = Canvas(path, pagesize=page_size)

        for pi, (page_start_idx, page_rows) in enumerate(pages):
            is_first = (pi == 0)

            if is_first:
                ty = ph - 8 * mm
                full_name = meta.get('full_name', '').strip()
                desc      = meta.get('description', '').strip()

                line1 = f"BOM Export \u2014 {meta.get('item_no', '')}"
                c.setFont('Helvetica-Bold', 11)
                c.setFillColor(colors.black)
                c.drawString(lm, ty, fit_text(line1, pw - lm - rm, 'Helvetica-Bold', 11))
                ty -= 5.5 * mm

                if desc or full_name:
                    if desc and full_name and desc != full_name:
                        line2 = f"{desc}  |  {full_name}"
                    else:
                        line2 = desc or full_name
                    c.setFont('Helvetica', 9)
                    c.setFillColor(colors.HexColor('#333333'))
                    c.drawString(lm, ty, fit_text(line2, pw - lm - rm, 'Helvetica', 9))
                    ty -= 5 * mm

                c.setFont('Helvetica', 8)
                c.setFillColor(colors.HexColor('#555555'))
                c.drawString(lm, ty,
                             f"Exported: {meta.get('exported_at', '')}  "
                             f"\u2022  Items: {meta.get('total_items', '')}")
                ty -= 4 * mm
                hdr_top = ty
            else:
                hdr_top = ph - 8 * mm

            # ── Blue header bar ───────────────────────────────────────
            hdr_bot = hdr_top - hdr_h
            c.setFillColor(hdr_bg_clr)
            c.rect(lm, hdr_bot, table_right - lm, hdr_h, fill=1, stroke=0)

            def draw_hdr_label(c, text, x, max_w, y_bot, bar_h, font, size):
                leading  = size * 1.2
                raw_words = text.replace('\u00a0', ' ').split()
                words = []
                for w in raw_words:
                    if '/' in w and stringWidth(w, font, size) > max_w - 4:
                        parts = w.split('/')
                        for i, p in enumerate(parts):
                            token = p + ('/' if i < len(parts) - 1 else '')
                            if token:
                                words.append(token)
                    else:
                        words.append(w)

                lines   = []
                current = ''
                for word in words:
                    test = (current + ' ' + word).strip()
                    if stringWidth(test, font, size) <= max_w - 4:
                        current = test
                    else:
                        if current:
                            lines.append(current)
                        current = word
                if current:
                    lines.append(current)
                lines = lines[:2]
                lines = [
                    fit_text(ln, max_w - 4, font, size)
                    if stringWidth(ln, font, size) > max_w - 4 else ln
                    for ln in lines
                ]

                total_h = len(lines) * leading
                y_start = y_bot + (bar_h + total_h) / 2 - leading
                c.setFont(font, size)
                for li, line in enumerate(lines):
                    c.drawString(x + 2, y_start - li * leading, line)

            c.setFillColor(colors.white)

            # Draw headers for all dynamic columns
            for i, (lbl, _, _) in enumerate(col_configs):
                draw_hdr_label(c, lbl, xs[i], cw_vals[i], hdr_bot, hdr_h, 'Helvetica-Bold', fs_hdr)

            # ── Row y-positions ───────────────────────────────────────
            # FIX 2: y_positions stays at geometric center (unchanged) so
            # connector dots remain aligned. Text is nudged separately below.
            y_top       = hdr_bot
            y_positions = [y_top - (i + 0.5) * row_h for i in range(len(page_rows))]
            centers     = y_positions[:]          # connectors use this directly
            row_tops    = [y_top - i * row_h      for i in range(len(page_rows))]

            # FIX 3: text baseline offset — moves text down from geometric
            # center so it appears visually centered (ReportLab baselines
            # sit above the visible text body).
            txt_offset = fs_body * 0.3

            # ── Data rows ─────────────────────────────────────────────
            for i, row in enumerate(page_rows):
                y     = y_positions[i]
                rt    = row_tops[i]
                depth = row['level']
                fnt   = 'Helvetica-Bold' if depth == 0 else 'Helvetica'

                # Depth-colored background
                draw_row_bg(c, rt, depth)

                # Horizontal grid line at row bottom
                draw_grid_line(c, lm, rt - row_h, table_right, rt - row_h, lw=0.3)

                c.setFillColor(colors.black)
                c.setFont(fnt, fs_body)

                # Draw data cells for all dynamic columns
                for i, (_, _, data_key) in enumerate(col_configs):
                    xc = xs[i]
                    cw = cw_vals[i]
                    if data_key == 'description':
                        # Designation cell — German (description) on top, English
                        # (full_name) below. Each line is gated by its own
                        # tree-header checkbox; lines also collapse when they
                        # would be empty or duplicate the other.
                        desc_de = str(row.get('description') or '').strip() if show_desc else ''
                        desc_en = str(row.get('full_name')   or '').strip() if show_name else ''
                        if desc_de and desc_en and desc_de != desc_en:
                            fs_en  = max(fs_body - 1, 5)
                            fnt_en = 'Helvetica-BoldOblique' if depth == 0 else 'Helvetica-Oblique'
                            c.setFont(fnt, fs_body)
                            c.setFillColor(colors.black)
                            c.drawString(xc + 2, y + row_h * 0.15,
                                         fit_text(desc_de, cw - conn_w - 4, fnt, fs_body))
                            c.setFont(fnt_en, fs_en)
                            c.setFillColor(colors.HexColor('#555555'))
                            c.drawString(xc + 2, y - row_h * 0.28,
                                         fit_text(desc_en, cw - conn_w - 4, fnt_en, fs_en))
                            c.setFillColor(colors.black)
                            c.setFont(fnt, fs_body)
                        else:
                            c.drawString(xc + 2, y - txt_offset,
                                         fit_text(desc_de or desc_en, cw - conn_w - 4, fnt, fs_body))
                    elif data_key == 'qty':
                        c.drawString(xc + 2, y - txt_offset,
                                     fit_text(_fmt_qty(row['qty']), cw - 4, fnt, fs_body))
                    elif data_key == 'position_path':
                        val = str(row.get('position_path') or row.get('position') or '')
                        c.drawString(xc + 2, y - txt_offset,
                                     fit_text(val, cw - 4, fnt, fs_body))
                    elif data_key == 'item_no':
                        c.drawString(xc + 2, y - txt_offset,
                                     fit_text(str(row['item_no']), cw - 4, fnt, fs_body))
                    elif data_key == 'scriptnum':
                        c.drawString(xc + 2, y - txt_offset,
                                     fit_text(str(row['scriptnum']), cw - 4, fnt, fs_body))
                    else:
                        c.drawString(xc + 2, y - txt_offset,
                                     fit_text(str(row.get(data_key, '')), cw - 4, fnt, fs_body))

            # ── Outer border + vertical dividers ─────────────────────
            if page_rows:
                table_bot = row_tops[-1] - row_h
            else:
                table_bot = hdr_bot

            c.setStrokeColor(grid_clr)
            c.setLineWidth(0.8)
            c.rect(lm, table_bot, table_right - lm, hdr_top - table_bot, fill=0, stroke=1)

            c.setLineWidth(0.3)
            # Draw vertical grid lines for all column dividers
            c.setLineWidth(0.3)
            for i in range(1, len(xs)):
                draw_grid_line(c, xs[i], hdr_top, xs[i], table_bot)

            # ── Connectors (inside Designation col right portion) ─────
            # Only render the tree connectors when the Designation column is
            # actually present — otherwise the spines would bleed into the
            # right-hand data cells.
            if page_rows and auto_idx >= 0:
                draw_connectors(c, page_rows, centers, page_start_idx)

            # ── Page number (right) + BOM number (left) ──────────────
            c.setFont('Helvetica', 7)
            c.setFillColorRGB(0.4, 0.4, 0.4)
            c.drawRightString(pw - rm, 4 * mm, f"Page {pi + 1} von {n_pages}")
            c.drawString(lm, 4 * mm, f"BOM: {meta.get('item_no', '')}")

            c.showPage()

        c.save()

    # ------------------------------------------------------------------ check logic
    def _on_item_check_changed(self, item: QTreeWidgetItem, column: int):
        """User clicked a checkbox — cascade to descendants / ancestors."""
        if column != 0:
            return
        if item.data(0, _ROLE_ITEM_NO) == _PLACEHOLDER:
            return

        self._tree.blockSignals(True)
        state = item.checkState(0)

        # Uncheck → propagate down to all loaded descendants
        # Check  → propagate down AND fix parent chain
        if state == Qt.CheckState.Unchecked:
            self._cascade_check(item, Qt.CheckState.Unchecked)
        else:
            self._cascade_check(item, Qt.CheckState.Checked)
            p = item.parent()
            while p is not None:
                p.setCheckState(0, Qt.CheckState.Checked)
                p = p.parent()

        self._tree.blockSignals(False)

    def _cascade_check(self, item: QTreeWidgetItem, state: Qt.CheckState):
        """Recursively apply *state* to all loaded (non-placeholder) descendants."""
        for i in range(item.childCount()):
            child = item.child(i)
            if child.data(0, _ROLE_ITEM_NO) == _PLACEHOLDER:
                continue
            child.setCheckState(0, state)
            self._cascade_check(child, state)

    # ------------------------------------------------------------------ misc
    def _clear(self):
        self._tree.clear()
        self._active_loaders.clear()
        self._export_pending = False
        self._status.setText("Cleared.")

    # ------------------------------------------------------------------ helpers
    def _start_loader(self, item_no: str, parent_tree_item: QTreeWidgetItem):
        dataset = self._dataset_cb.currentText()
        loader  = BOMLoader(item_no, dataset)
        loader.data_ready.connect(
            lambda rows, p=parent_tree_item: self._on_data_ready(p, rows)
        )
        loader.error.connect(self._on_error)
        loader.finished.connect(lambda l=loader: self._cleanup_loader(l))
        loader.start()
        self._active_loaders.append(loader)

    def _cleanup_loader(self, loader: BOMLoader):
        try:
            self._active_loaders.remove(loader)
        except ValueError:
            pass
        if self._export_pending and not self._active_loaders:
            self._export_pending = False
            self._export_bom()

    def _make_node(self, parent, pos, item_no, qty,
                   has_bom, description, full_name, scriptnum='',
                   stockloc='', bestand='') -> QTreeWidgetItem:
        """Create and return a properly configured QTreeWidgetItem."""
        node = _BOMTreeItem(parent)
        node.setText(0, pos)
        node.setText(1, item_no)
        node.setText(2, qty)
        node.setText(3, str(scriptnum) if scriptnum else '')
        node.setText(4, description)
        node.setText(5, full_name)
        node.setText(6, str(stockloc) if stockloc else '')
        node.setText(7, _fmt_qty(bestand) if bestand != '' else '')
        node.setData(0, _ROLE_ITEM_NO, item_no)
        node.setData(0, _ROLE_HAS_BOM, has_bom)
        node.setFlags(node.flags() | Qt.ItemFlag.ItemIsUserCheckable)
        node.setCheckState(0, Qt.CheckState.Checked)

        # ── Depth-based gray shading ──────────────────────────────────────
        depth = 0
        p = node.parent()
        while p is not None:
            depth += 1
            p = p.parent()
        bg = QColor(_DEPTH_COLORS[min(depth, len(_DEPTH_COLORS) - 1)])
        for col in range(self._tree.columnCount()):
            node.setBackground(col, bg)

        if has_bom:
            ph = _BOMTreeItem(node)
            ph.setText(1, '...')
            ph.setData(0, _ROLE_ITEM_NO, _PLACEHOLDER)

        return node